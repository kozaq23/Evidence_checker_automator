# 2025-07-07 13.41.49
# ECA = Evidence Checker Automation

import sys
import uuid
import platform
import socket
import time
import logging
import tkinter as tk
from tkinter import filedialog, ttk
import tkinter.font as tkfont
from pathlib import Path
import logging
from dataclasses import dataclass
from typing import List, Tuple, Optional, Dict
import re
import os

import pandas as pd # Pandas >= 1.2.0 
import openpyxl # and Openpyxl >= 3.0.0.
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries

logging.basicConfig(filename="my_log_file.log", level=logging.DEBUG, format="%(asctime)s %(levelname)s:%(message)s")
logger = logging.getLogger(__name__)
check_counts: Dict[str, int] = {}

EXPECTED_TOTAL_SALES = 7_777_460_207

def log_startup(config):
    start_ts = time.time()
    run_id = uuid.uuid4()
    logger.info("="*40)  # unique identifier
    logger.info(f"Run ID: {run_id}")  # unique identifier
    logger.info(f"Startup timestamp: {start_ts}")  # epoch start time
    # Python and dependency versions
    logger.info(f"Python version: {platform.python_version()}")
    logger.info(f"pandas version: {pd.__version__}")
    logger.info(f"openpyxl version: {openpyxl.__version__}")
    # OS and hostname
    logger.info(f"Hostname: {socket.gethostname()}")
    # Configuration settings and selection method
    logger.info(
        f"Config: dark_mode={config.dark_mode}, choice={config.choice}, "
        f"show_info={config.show_info}, show_ok={config.show_ok}, show_errors={config.show_errors}")
    return start_ts, run_id

def log_shutdown(start_ts, run_id, path, messages, exit_code=0):
    end_ts = time.time()
    duration = end_ts - start_ts
    # File info and type detected
    logger.info(f"Run ID: {run_id} - analysed file: {path}")
    # Count message levels
    counts = {'info': sum(1 for _,lvl in messages if lvl=='info'),
              'ok': sum(1 for _,lvl in messages if lvl=='ok'),
              'error': sum(1 for _,lvl in messages if lvl=='error') }
    logger.info(f"Message counts: {counts}")
    # Detailed failed checks
    failed = [text for text, lvl in messages if lvl=='error']
    logger.info(f"Failed checks ({len(failed)}): {failed}")
    # Check invocation counts
    logger.info(f"Check function invocation counts: {check_counts}")
    # Execution duration and exit status
    logger.info(f"Execution duration: {duration} seconds")
    logger.info(f"Exit status code: {exit_code}")
    return exit_code

def identify_wp3_file(path: Path) -> str:
    check_counts['identify_wp3_file'] = check_counts.get('identify_wp3_file', 0) + 1
    try:
        df = pd.read_excel(path)
        rows, _ = df.shape
        columns = set(df.columns)
        if rows > 400 and columns == {"Year", "Album", "Artist", "Total Sales"}: # 400+ lines and exact order of columns
            return "music" # flag it as music dataset exercise
        required_columns = {"Name", "Date", "Department", "Rating"}
        if rows > 90 and required_columns.issubset(set(columns)): # 90 + lines and contains minimum of column names listed line above
            return "dashboard" # flag it as dashboard exercise
    except Exception as e:
        return "error" # file probably open with Excel
        messages.append((f"Cannot open file: {e}", "error"))
        if "[Errno 13] Permission denied" in str(e):
            return messages

@dataclass
class Config:
    last_dir: Optional[str] = None
    manual_dir: Optional[str] = None
    dark_mode: bool = False
    choice: str = "latest"
    geometry: Optional[str] = None
    show_info: bool = True
    show_ok: bool = True
    show_errors: bool = True

    @classmethod
    def load(cls, path: Path) -> "Config":
        if not path.exists():
            return cls(dark_mode=True)
        try:
            data = {}
            for line in path.read_text().splitlines():
                if not line.strip():
                    continue
                key, val = line.split("=", 1)
                data[key] = val
            return cls(last_dir=data.get("last_dir"),
                manual_dir=data.get("manual_dir"),
                dark_mode=data.get("dark_mode", "false").lower() == "true",
                choice=data.get("choice") if data.get("choice") in ("latest", "manual") else "latest",
                geometry=data.get("geometry"),
                show_info=data.get("show_info", "true").lower() == "true",
                show_ok=data.get("show_ok", "true").lower() == "true",
                show_errors=data.get("show_errors", "true").lower() == "true")
        except Exception as e:
            logging.error(f"Error loading config: {e}")
            return cls(dark_mode=True)

    def save(self, path: Path) -> None:
        entries = []
        if self.last_dir:
            entries.append(f"last_dir={self.last_dir}")
        if self.manual_dir:
            entries.append(f"manual_dir={self.manual_dir}")
        entries.append(f"dark_mode={'true' if self.dark_mode else 'false'}")
        entries.append(f"choice={self.choice}")
        if self.geometry:
            entries.append(f"geometry={self.geometry}")
        entries.append(f"show_info={'true' if self.show_info else 'false'}")
        entries.append(f"show_ok={'true' if self.show_ok else 'false'}")
        entries.append(f"show_errors={'true' if self.show_errors else 'false'}")
        path.write_text("\n".join(entries))

config = Config.load(Path("config_ECA.txt"))
current_config = config

def find_latest_excel(download_folder: Path) -> Optional[Path]:
    if not download_folder.is_dir():
        logging.warning("Downloads folder not found")
        return None
    files = [f for f in download_folder.iterdir() if f.suffix.lower() in (".xls", ".xlsx") and f.is_file()] # finds all excel files
    if not files:
        logging.info("No Excel files in Downloads")
        return None
    return max(files, key=lambda f: f.stat().st_mtime) # highest time of last modification [which is the latest modified file]

def select_appropriate_sheet(excel: pd.ExcelFile) -> Tuple[Optional[pd.DataFrame], Optional[str], List[Tuple[str, str]]]:
    messages: List[Tuple[str, str]] = []
    sheets = excel.sheet_names
    if len(sheets) == 1 and sheets[0].strip().upper() == "RAW DATA":
        messages.append(("[Warning] Only 'RAW DATA' found.", "info"))
        try:
            df = excel.parse(sheets[0])
            return df, sheets[0], messages
        except Exception as e:
            messages.append((f"Cannot parse 'RAW DATA': {e}", "error"))
            return None, None, messages
    messages.append((f"Workbook sheet(s) detected: {sheets}", "info"))
    
    for sheet in sheets:
        if sheet.strip().upper() == "RAW DATA":
            continue
        try:
            df_candidate = excel.parse(sheet)
        except Exception:
            continue
        if df_candidate.shape[1] < 1:
            continue
        last_col = df_candidate.iloc[:, -1]
        if pd.api.types.is_numeric_dtype(last_col) and last_col.count() >= 400: # if the last column is numerical and has more than 400 rows select it.
            messages.append((f"Reviewing the sheet '{sheet}'.", "info"))
            return df_candidate, sheet, messages
            
    if "RAW DATA" in (s.upper() for s in sheets):
        messages.append(("No suitable sheet found - falling back to 'RAW DATA'.", "info"))
        try:
            df = excel.parse("RAW DATA")
            return df, "RAW DATA", messages
        except Exception as e:
            messages.append((f"I couldn't open the 'RAW DATA' sheet ({e}).", "error"))
            return None, None, messages
    messages.append(("The workbook has no 'RAW DATA' sheet.", "error"))
    return None, None, messages

def auto_select_sheet(excel: pd.ExcelFile) -> Tuple[Optional[pd.DataFrame], Optional[str], List[Tuple[str, str]]]:
    messages: List[Tuple[str, str]] = []
    sheets = excel.sheet_names
    if len(sheets) == 1 and sheets[0].strip().upper() == "TASK ONE":
        messages.append(("'TASK ONE' found.", "info"))
        try:
            df = excel.parse(sheets[0])
            return df, sheets[0], messages
        except Exception as e:
            messages.append((f"Cannot parse 'Task One': {e}", "error"))
            return None, None, messages
        messages.append((f"Workbook sheet(s) detected: {sheets}", "info"))
    elif len(sheets) == 1:
        messages.append((f"{sheets} sheet found, and it is the only one in this file, proceeding.", "info"))
        try:
            df = excel.parse(sheets[0])
            return df, sheets[0], messages
        except Exception as e:
            messages.append((f"Cannot parse '{sheets}': {e}", "error"))
            return None, None, messages
            
    for sheet in sheets:
        if sheet.strip().upper() == "TASK ONE":
            continue
        try:
            df_candidate = excel.parse(sheet)
        except Exception:
            continue
        if df_candidate.shape[1] < 1:
            continue
        last_col = df_candidate['Rating']
        if last_col.count() > 90: # if the Rating column is numerical and has more than 90 rows select it.
            messages.append((f"Reviewing the sheet '{sheet}'.", "info"))
            return df_candidate, sheet, messages
        
def check_nulls(df: pd.DataFrame) -> List[Tuple[str, str]]:
    check_counts['check_nulls'] = check_counts.get('check_nulls', 0) + 1
    results: List[Tuple[str, str]] = []
    if df.isnull().any().any():
        results.append(("Some cells are blank.", "error"))
        rows = [i + 2 for i in df[df.isnull().any(axis=1)].index]
        results.append((f"Blank cells found in rows: {rows}.", "error"))
    else:
        results.append(("No blank cells found. [OK]", "ok"))
    return results

def check_artist_column(df: pd.DataFrame) -> List[Tuple[str, str]]:
    check_counts['check_artist_column'] = check_counts.get('check_artist_column', 0) + 1
    results: List[Tuple[str, str]] = []
    if "Artist" not in df.columns:
        return results
    trim_errors: List[Tuple[int, str, str]] = []
    case_errors: List[Tuple[int, str, str]] = []
    for idx, val in enumerate(df["Artist"].astype(str)):
        trimmed = val.strip()
        proper = trimmed.title()
        if val != trimmed:
            trim_errors.append((idx + 2, val, trimmed))
        if trimmed != proper:
            case_errors.append((idx + 2, trimmed, proper))
    total = len(trim_errors) + len(case_errors)
    if total:
        results.append((f"'Artist' column needs {total} corrections.", "error"))
        for r, fnd, exp in trim_errors:
            results.append((f"Row {r} 'Artist': remove extra spaces ('{fnd}' → '{exp}').", "error"))
        for r, fnd, exp in case_errors:
            results.append((f"'Row {r} 'Artist': adjust capitalisation ('{fnd}' → '{exp}').", "error"))
    else:
        results.append(("'Artist' entries are trimmed and capitalised correctly.", "ok"))
    return results

def check_duplicates(df: pd.DataFrame) -> List[Tuple[str, str]]:
    check_counts['check_duplicates'] = check_counts.get('check_duplicates', 0) + 1
    results: List[Tuple[str, str]] = []
    duplicates = df.duplicated(keep=False)
    if duplicates.any():
        results.append(("Duplicate rows detected.", "error"))
        indices = df[duplicates].index.tolist()
        results.append((f"Duplicate row numbers: {indices}.", "error"))
    else:
        results.append(("No duplicate rows detected. [OK]", "ok"))
    return results

def check_album_duplicates(df: pd.DataFrame) -> List[Tuple[str, str]]:
    check_counts['check_album_duplicates'] = check_counts.get('check_album_duplicates', 0) + 1    
    results: List[Tuple[str, str]] = []
    if "Album" not in df.columns:
        return results
    count = (df["Album"].astype(str).str.lower() == "greatest hits").sum()
    if count > 1:
        results.append(("Multiple 'Greatest Hits' entries found in 'Album' - expected. [OK]", "ok"))
    else:
        results.append(("'Greatest Hits' does not appear more than once in 'Album'.", "error"))
    return results

def check_total_sales(df: pd.DataFrame, path: Path, sheet: str) -> List[Tuple[str, str]]:
    check_counts['check_total_sales'] = check_counts.get('check_total_sales', 0) + 1        
    results: List[Tuple[str, str]] = []
    cols = list(df.columns)
    if "Total Sales" in cols:
        col_name = "Total Sales"
        total = df[col_name].sum()
        if total != EXPECTED_TOTAL_SALES:
            results.append((f"Total Sales sum is {total}; expected {EXPECTED_TOTAL_SALES}.", "error"))
        else:
            results.append((f"'{col_name}' total matches the expected figure. [OK]", "ok"))
    else:
        col_name = cols[-1]
        results.append((f"'Total Sales' column missing - using '{col_name}' instead.", "info"))
    try:
        wb = openpyxl.load_workbook(str(path), data_only=False)
        ws = wb[sheet]
        idx = cols.index(col_name) + 1
        letter = get_column_letter(idx)
        fmt_ok = True
        prec_ok = True
        for r in range(2, df.shape[0] + 2):
            nf = str(ws[f"{letter}{r}"].number_format)
            if "£" not in nf: fmt_ok = False
            if not re.search(r"0\.00", nf): prec_ok = False
            if not fmt_ok or not prec_ok: break
        if fmt_ok:
            results.append((f"'{col_name}' is formatted as GBP Accounting. [OK]", "ok"))
        else:
            results.append((f"'{col_name}' is not formatted as GBP Accounting.", "error"))
        if prec_ok:
            results.append((f"'{col_name}' shows two decimal places. [OK]", "ok"))
        else:
            results.append((f"'{col_name}' does not show two decimal places.", "error"))
    except Exception as e:
        results.append((f"Couldn't verify the format or precision of '{col_name}' ({e}).", "error"))
    return results

def check_qs(df: pd.DataFrame, path: Path, sheet: str) -> List[Tuple[str, str]]:
    check_counts['check_qs'] = check_counts.get('check_qs', 0) + 1            
    results: List[Tuple[str, str]] = []
    cols = list(df.columns)
    
    text_series = df.astype(str).apply(lambda x: ' '.join(x), axis=1)
    qs_count = text_series.str.count(r'\bQS\b', flags=re.IGNORECASE).sum()
    quality_surveyor_count = text_series.str.count(r'\bQuality Surveyor\b', flags=re.IGNORECASE).sum()
    misspellings = text_series[text_series.str.contains(r'\bq\w*\s+s\w*', flags=re.IGNORECASE)]
    if qs_count > 1:
        results.append((f"QS not changed to Quality Surveyor, found QS {qs_count} times.", "error"))
    elif len(misspellings) > 0:
        results.append((f"QS changed to {set(misspellings)}.", "error"))
    elif quality_surveyor_count >= 16:
        results.append((f"QS changed to Quality Surveyor. [OK]", "ok"))
    return results

def check_validation(df: pd.DataFrame, path: Path, sheet: str) -> List[Tuple[str, str]]:
    check_counts['check_validation'] = check_counts.get('check_validation', 0) + 1
    results: List[Tuple[str, str]] = []
    cols = list(df.columns)
    wb = load_workbook(path)
    ws = wb[sheet]
    # Identify the target column by header name
    header_row = 1
    # Map headers to column indices
    header_to_col = {cell.value: cell.column for cell in ws[header_row]}
    # Check each validation rule
    def validate(type_of_validation, column_name):
        target_header = column_name
        target_col_index = header_to_col[column_name]
        target_col_letter = get_column_letter(target_col_index)
        # Get all data validations
        validations = list(ws.data_validations.dataValidation)
        applied = False
        for dv in validations:
            if dv.type != type_of_validation:
                continue
            for cell_range in dv.sqref:  # each item is already a CellRange
                min_col, min_row, max_col, max_row = cell_range.bounds
            if min_col <= target_col_index <= max_col:
                applied = True
                results.append((f"'{type_of_validation}' validation applied to '{target_header}' in range {cell_range} [OK]", "ok"))
                break
            if applied:
                break
        if not applied:
            results.append((f"No '{type_of_validation}' data validation found for column '{target_header}'.", "error"))
    validate(type_of_validation='whole', column_name='Rating')
    validate(type_of_validation='list', column_name='Department')
    return results

def check_functions(df: pd.DataFrame, path: Path, sheet: str) -> List[Tuple[str, str]]:
    check_counts['check_functions'] = check_counts.get('check_functions', 0) + 1
    results: List[Tuple[str, str]] = []

    primary_functions = {'SUM', 'MAX', 'MIN', 'AVERAGE', 'MEDIAN', 'MODE', 'STDEV.S'}
    alternative_forms = {'STDEV.S': 'STDEV'}

    found_functions = set()
    found_alternatives = set()

    wb = load_workbook(path, data_only=False)
    ws = wb[sheet]

    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f' and cell.value:
                formula = str(cell.value).upper()
                for func in primary_functions:
                    if func in formula:
                        found_functions.add(func)
                for alt_func in alternative_forms.values():
                    if alt_func in formula:
                        found_alternatives.add(alt_func)
    missing_functions = primary_functions - found_functions

    # Check for acceptable alternatives
    for primary, alt in alternative_forms.items():
        if primary in missing_functions and alt in found_alternatives:
            results.append((f"{primary} not found, but alternative function {alt} has been used. [OK]", "ok"))
            missing_functions.discard(primary)

    if not missing_functions:
        results.append(("All required functions found/applied. [OK]", "ok"))
    elif found_functions or found_alternatives:
        found_list = ', '.join(sorted(found_functions.union(found_alternatives)))
        missing_list = ', '.join(sorted(missing_functions))
        results.append((f"Functions found: {found_list}", "ok"))
        results.append((f"Functions missing: {missing_list}", "error"))
    else:
        results.append(("No required functions found/applied in the spreadsheet.", "error"))
    return results

def check_table_format(df: pd.DataFrame, path: Path, sheet: str) -> List[Tuple[str, str]]:
    check_counts['check_table_format'] = check_counts.get('check_table_format', 0) + 1
    results: List[Tuple[str, str]] = []
    try:
        wb = openpyxl.load_workbook(str(path), data_only=False)
        ws = wb[sheet]
        tables = ws.tables
        if not tables:
            results.append(("The data isn't in an Excel table format.", "error"))
            return results
        if len(tables) > 1:
            results.append(("More than one Excel table found on the sheet.", "error"))
            return results
        tbl = next(iter(tables.values()))
        min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
        expected_max_col = df.shape[1]
        expected_max_row = df.shape[0] + 1
        if (min_row, min_col) == (1, 1) and max_col == expected_max_col:
            if max_row in (expected_max_row, 1_048_573):
                results.append(("The table range fits the data exactly. [OK]", "ok"))
            else:
                results.append(("The table range doesn't match the data exactly.", "error"))
        else:
            results.append(("The table range doesn't match the data exactly.", "error"))
    except Exception as e:
        results.append((f"Cannot verify table format: {e}", "error"))
    return results

def analyse_excel(path: Path) -> List[Tuple[str, str]]:
    messages: List[Tuple[str, str]] = []
    file_type = identify_wp3_file(path)

    if file_type == "music":
        messages.append(("Detected WP3 - Music Data", "info"))
        excel_file = pd.ExcelFile(str(path))

        df, sheet, sel_msgs = select_appropriate_sheet(excel_file)
        messages.extend(sel_msgs)
        if df is None or sheet is None:
            return messages
        messages.extend(check_nulls(df))
        messages.extend(check_artist_column(df))
        messages.extend(check_duplicates(df))
        messages.extend(check_album_duplicates(df))
        messages.extend(check_total_sales(df, path, sheet))
        messages.extend(check_table_format(df, path, sheet))
        return messages
        
    elif file_type == "dashboard":
        messages.append(("Detected WP3 - Excel Stats Dashboard", "info"))
        excel_file = pd.ExcelFile(str(path))
        df, sheet, sel_msgs = auto_select_sheet(excel_file)
        messages.extend(sel_msgs)
        if df is None or sheet is None:
            return messages
        # to finalise:
        # check if QS is Quality Surveyor [outcomes[+/-]]
        messages.extend(check_qs(df, path, sheet))
        
        # check if Data Validation is applied [outcomes[+/-/partial(granual)]]
        messages.extend(check_validation(df, path, sheet))
        
        # check if functions like =SUM(), =MAX(), =MIN(), =AVERAGE(), =MEDIAN(), =MODE(), =STDEV.S() are used in the spreadsheet [outcomes[+/-/partial]].
        messages.extend(check_functions(df, path, sheet))
        return messages
    elif file_type == "error":
        messages.append(("Close Excel with the workbook and run the check again.", "error"))
        return messages
    else:
        messages.append(("File did not match any known WP3 format", "error"))
        return messages


class ToolTip:
    def __init__(self, widget: tk.Widget, text_fn):
        self.widget = widget
        self.text_fn = text_fn
        self.tip_window = None
        widget.bind("<Enter>", self.show_tip)
        widget.bind("<Leave>", self.hide_tip)

    def show_tip(self, _=None):
        if self.tip_window:
            return
        content = self.text_fn()
        if not content:
            return
        bg = "#333333" if current_config.dark_mode else "#ffffe0"
        fg = "#ffffff" if current_config.dark_mode else "#000000"
        x, y, *_ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 20
        y += self.widget.winfo_rooty() + 20
        self.tip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(
            tw,
            text=content,
            justify="left",
            background=bg,
            foreground=fg,
            relief="solid",
            borderwidth=1
        )
        label.pack(ipadx=4, ipady=2)

    def hide_tip(self, _=None):
        if self.tip_window:
            self.tip_window.destroy()
            self.tip_window = None
            

class EvidenceCheckerUI:
    def __init__(self, root: tk.Tk, config: Config, config_path: Path):
        self.root = root
        self.config = config
        self.config_path = config_path
        self.analysis_messages: List[Tuple[str, str]] = []
        self.status_job = None
        self.default_font = tkfont.nametofont("TkDefaultFont")
        self.bold_font    = self.default_font.copy()
        self.bold_font.configure(weight="bold")
        self.root.title("Evidence Checking Automator [WP3]")
        self._init_variables()
        self._load_geometry()
        self._build_ui()
        self._apply_dark_mode()
    
    def _init_variables(self):
        self.show_info = tk.BooleanVar(value=self.config.show_info)
        self.show_ok = tk.BooleanVar(value=self.config.show_ok)
        self.show_err = tk.BooleanVar(value=self.config.show_errors)
        self.choice_var = tk.StringVar(value=self.config.choice)
        self.path_var = tk.StringVar()
        self.status_var = tk.StringVar(value="Ready")

    def _load_geometry(self):
        if self.config.geometry:
            self.root.minsize(1000, 480)
            self.root.geometry(self.config.geometry)
        else:
            self.root.minsize(1000, 480)
            
    def _save_config(self):
        self.config.last_dir = self.config.last_dir
        self.config.manual_dir = self.config.manual_dir
        self.config.dark_mode = self.config.dark_mode
        self.config.choice = self.choice_var.get()
        self.config.geometry = self.root.geometry()
        self.config.show_info = self.show_info.get()
        self.config.show_ok = self.show_ok.get()
        self.config.show_errors = self.show_err.get()
        self.config.save(self.config_path)

    def _apply_dark_mode(self):
        bg = "#2e2e2e" if self.config.dark_mode else "#f0f0f0"
        fg = "#ffffff" if self.config.dark_mode else "#000000"
        self.root.configure(bg=bg)
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TFrame", background=bg)
        style.configure("TLabel", background=bg, foreground=fg)
        style.configure("TButton", background=bg, foreground=fg)
        style.map("TButton", background=[("active", bg)], foreground=[("active", fg)])
        style.map("TRadiobutton", background=[("active", "#555555" if self.config.dark_mode else "#d0d0d0")], foreground=[("active", fg)])
        style.map("TCheckbutton", background=[("active", "#555555" if self.config.dark_mode else "#d0d0d0")], foreground=[("active", fg)])
        style.configure("TRadiobutton", background=bg, foreground=fg)
        style.configure("TCheckbutton", background=bg, foreground=fg)
        style.configure("TEntry", fieldbackground="#3e3e3e" if self.config.dark_mode else "#ffffff", foreground=fg)
        style.configure("Treeview", background="#3e3e3e" if self.config.dark_mode else "#ffffff", foreground=fg, fieldbackground="#3e3e3e" if self.config.dark_mode else "#ffffff")
        style.configure("TScrollbar", troughcolor=bg, background="#4e4e4e" if self.config.dark_mode else "#d0d0d0")
        style.map("TScrollbar", background=[("active", "#666666" if self.config.dark_mode else "#b0b0b0")])
        style.configure("Status.TLabel", background=bg, foreground=fg)

    def _build_ui(self):
        for w in self.root.winfo_children():
            w.destroy()
        bg = "#2e2e2e" if self.config.dark_mode else "#f0f0f0"
        fg = "#ffffff" if self.config.dark_mode else "#000000"
        top_frame = ttk.Frame(self.root)
        top_frame.pack(fill="x", padx=5, pady=1)
        side_frame = ttk.Frame(top_frame)
        side_frame.pack(side="right")

        btn = ttk.Button(side_frame,
        text="☀" if self.config.dark_mode else "🌙",
        command=self._toggle_dark_mode)
        btn.pack(side="left", padx=1)
        ToolTip(btn, lambda: "Switch to bright mode" if self.config.dark_mode else "Switch to dark mode")

        rb_latest = ttk.Radiobutton(top_frame, text="Select latest Excel in Downloads", variable=self.choice_var, value="latest")
        rb_latest.pack(side="left", padx=0)
        ToolTip(rb_latest, lambda: "Auto-select most recent Excel")
        rb_manual = ttk.Radiobutton(self.root, text="Pick Excel manually", variable=self.choice_var, value="manual")
        rb_manual.pack(anchor="w", padx=5, pady=1)
        ToolTip(rb_manual, lambda: "Choose an Excel file")

        self.select_btn = ttk.Button(self.root, text="Auto-select [Downloads]" if self.choice_var.get()=="latest" else "Pick file…",
                                     command=self._update_path)
        self.select_btn.pack(pady=1)
        ToolTip(self.select_btn, lambda: "Click to select Excel file")

        self.path_disp = tk.Text(self.root, height=1, font=self.default_font, 
                                 bg="#3e3e3e" if self.config.dark_mode else "#ffffff",
                                 fg="#ffffff" if self.config.dark_mode else "#000000",
                                 wrap="none",
                                 bd=0,
                                 highlightthickness=0)
        self.path_disp.tag_config("bold", font=self.bold_font)
        self.path_disp.config(state="disabled")
        self.path_disp.pack(padx=5, pady=2, fill="x")
        
        ToolTip(self.path_disp, lambda: "Shows the full path")

        self.path_var.trace_add("write", self._refresh_path_disp)
        self._refresh_path_disp()

        self.analyse_btn = ttk.Button(self.root, text="Analyse", command=self._analyse_file)
        self.analyse_btn.pack(pady=5)
        ToolTip(self.analyse_btn, lambda: "Run analysis")

        filter_frame = ttk.Frame(self.root)
        filter_frame.pack(fill="x", padx=5, pady=1)
        for text, var in (("Show info", self.show_info), ("Show OK", self.show_ok), ("Show errors", self.show_err)):
            cb = ttk.Checkbutton(filter_frame, text=text, variable=var, command=lambda: (self._save_config(), self._display_analysis()))
            cb.pack(side="left", expand=True, fill="x")
        ToolTip(filter_frame, lambda: "Toggle messages")

        self.path_var.trace_add("write", lambda *_: self._toggle_analyse_btn())
        self._toggle_analyse_btn()

        output_frame = ttk.Frame(self.root)
        output_frame.pack(fill="both", expand=True, padx=5, pady=5)
        self.analysis_output = tk.Text(output_frame, height=10,
                                       bg="#222222" if self.config.dark_mode else "#ffffff",
                                       fg=fg, wrap="none", bd=0, highlightthickness=0)
        self.analysis_output.bind("<Key>", lambda e: "break")
        self.analysis_output.bind("<Control-c>", lambda e: self.analysis_output.event_generate("<<Copy>>"))
        v_scroll = ttk.Scrollbar(output_frame, orient="vertical", command=self.analysis_output.yview)
        self.analysis_output.config(yscrollcommand=v_scroll.set)
        self.analysis_output.pack(side="left", fill="both", expand=True)
        v_scroll.pack(side="right", fill="y")
        if self.analysis_messages:
            self._display_analysis()

        self.status_bar = ttk.Label(self.root, textvariable=self.status_var, anchor="w", relief="sunken", style="Status.TLabel")
        self.status_bar.pack(side="bottom", fill="x")
        ToolTip(self.status_bar, lambda: "Status messages")
        self.choice_var.trace_add("write", lambda *_: self._update_select_btn_text())

    def _set_status(self, message: str):
        self.status_var.set(message)
        if self.status_job:
            self.root.after_cancel(self.status_job)
        self.status_job = self.root.after(3000, lambda: self.status_var.set("Ready"))

    def _update_select_btn_text(self):
        text = "Auto-select [Downloads]" if self.choice_var.get()=="latest" else "Pick file…"
        self.select_btn.config(text=text)

    def _refresh_path_disp(self, *_):
        self.path_disp.config(state="normal")
        self.path_disp.delete("1.0", tk.END)
        p = self.path_var.get()
        if p:
            d, f = Path(p).parent, Path(p).name
            self.path_disp.insert("1.0", str(d) + os.sep)
            self.path_disp.insert("end", f, "bold")
        else:
            self.path_disp.insert("1.0", "Path will appear here")
        self.path_disp.config(state="disabled")

    def _toggle_analyse_btn(self):
        p = self.path_var.get().strip()
        valid = Path(p).is_file() and p.lower().endswith((".xls", ".xlsx"))
        self.analyse_btn.config(state="normal" if valid else "disabled")

    def _update_path(self):
        if self.choice_var.get()=="latest":
            latest = find_latest_excel(Path.home()/"Downloads")
            if latest:
                self.path_var.set(str(latest))
                self.config.last_dir = str(latest.parent)
                self._save_config()
            else:
                self.path_var.set("No Excel files in Downloads")
        else:
            initial = Path(self.config.manual_dir) if self.config.manual_dir else Path.home()
            sel = filedialog.askopenfilename(initialdir=str(initial), filetypes=[("Excel files","*.xls *.xlsx")])
            if sel:
                self.path_var.set(sel)
                self.config.manual_dir = str(Path(sel).parent)
                self.config.last_dir = str(Path(sel).parent)
                self._save_config()

    def _toggle_dark_mode(self):
        self.config.dark_mode = not self.config.dark_mode
        self._save_config()
        self._apply_dark_mode()
        self._build_ui()
        self._set_status("Dark mode toggled")

    def _analyse_file(self):
        path = Path(self.path_var.get().strip())
        self.analysis_messages = analyse_excel(path)
        self._display_analysis()
        self._set_status("Analysis done")

    def _display_analysis(self):
        self.analysis_output.config(state="normal")
        self.analysis_output.delete("1.0", tk.END)
        styles = self._get_styles()
        for text, level in self.analysis_messages:
            if level=="info" and not self.show_info.get(): continue
            if level=="ok" and not self.show_ok.get(): continue
            if level=="error" and not self.show_err.get(): continue
            self.analysis_output.insert("end", text+"\n", level)
        for tag, cfg in styles.items():
            self.analysis_output.tag_config(tag, foreground=cfg["fg"], background=cfg["bg"],
                                            selectbackground=cfg["sel_bg"], selectforeground=cfg["sel_fg"])
        self.analysis_output.config(state="disabled")

    def _get_styles(self) -> Dict[str, Dict[str, str]]:
        if self.config.dark_mode:
            return {"info": {"fg":"#00ffff","bg":"#222222","sel_bg":"#555555","sel_fg":"#00ffff"},
                    "ok":   {"fg":"#00ff00","bg":"#222222","sel_bg":"#555555","sel_fg":"#00ff00"},
                    "error":{"fg":"#ff0000","bg":"#222222","sel_bg":"#555555","sel_fg":"#ff0000"}}
        return {"info": {"fg":"blue","bg":"#ffffff","sel_bg":"#cce6ff","sel_fg":"blue"},
                "ok":   {"fg":"green","bg":"#ffffff","sel_bg":"#cce6ff","sel_fg":"green"},
                "error":{"fg":"red","bg":"#ffffff","sel_bg":"#cce6ff","sel_fg":"red"}}

    def run(self):
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        self.root.mainloop()

    def _on_close(self):
        self._save_config()
        self.root.destroy()

if __name__=="__main__":
    if len(sys.argv) > 1 and sys.argv[1] == 'test':
        import unittest

        class TestEvidenceChecker(unittest.TestCase):
            def test_identify_unknown(self):
                # Path to a non-existent file triggers unknown
                self.assertEqual(identify_wp3_file(Path("nonexistent.xlsx")), "error")

            def test_check_nulls_ok(self):
                df = pd.DataFrame({'A': [1,2], 'B': [3,4]})
                msgs = check_nulls(df)
                self.assertIn(("No blank cells found. [OK]", "ok"), msgs)

        unittest.main(argv=['first-arg-is-ignored'], exit=False)
    else:
        # Application run
        config = Config.load(Path("config_ECA.txt"))
        start_ts, run_id = log_startup(config)
        exit_code = 0
        try:
            root = tk.Tk()
            app = EvidenceCheckerUI(root, config, Path("config_ECA.txt"))
            app.run()
        except Exception as e:
            logger.exception(f"Unexpected error: {e}")
            exit_code = 1
        finally:
            # Path and messages may not exist if early failure; guard defaults
            path = getattr(app, 'path_var', '')
            messages = getattr(app, 'analysis_messages', [])
            code = log_shutdown(start_ts, run_id, path, messages, exit_code)
            # sys.exit(code)
